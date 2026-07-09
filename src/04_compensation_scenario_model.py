"""
04_compensation_scenario_model.py
=================================
COMPENSATION COST & SCENARIO MODEL

Projects total workforce cost over a 24-month horizon under three scenarios, and
exports a fully-formula-driven Excel workbook a finance team can actually use.

SCENARIOS
---------
A) Status Quo        — no intervention; high-risk employees churn at their modelled
                       rate and are replaced at full replacement cost.
B) Proactive Retention — give targeted retention raises to watchlist employees,
                       which lowers their churn probability; you pay more salary but
                       avoid most replacement cost.
C) Reactive / Do-Nothing — same as status quo but assumes the business reacts only
                       after exits, eating the full replacement cost with no salary
                       savings and a productivity gap.

REPLACEMENT COST MODEL
----------------------
Replacement cost is expressed as a multiplier of annual base salary, scaled by role
seniority (junior roles cheaper to replace, leaders far more expensive). The
multiplier bundles recruiting fees, onboarding/ramp, manager time, and lost
productivity — consistent with widely-cited SHRM/industry ranges (50%-200%).

    L1 Junior   : 0.50x
    L2 Mid      : 0.75x
    L3 Senior   : 1.00x
    L4 Lead     : 1.25x
    L5 Manager  : 1.50x
    L6 Director : 2.00x

The model is intentionally transparent: every assumption is a named input so a
technical interviewer can change a number and watch the whole thing recompute.

Outputs:
    outputs/scenario_summary.csv             tidy scenario results
    outputs/compensation_scenario_model.xlsx Excel model with live formulas
    (returns a DataFrame for the dashboard)

Run:  python src/04_compensation_scenario_model.py
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"

# ---------------------------------------------------------------------------
# Assumptions (single source of truth — also written into the Excel inputs tab)
# ---------------------------------------------------------------------------
HORIZON_MONTHS = 24
REPLACEMENT_MULTIPLIER = {
    "L1 - Junior": 0.50, "L2 - Mid": 0.75, "L3 - Senior": 1.00,
    "L4 - Lead": 1.25, "L5 - Manager": 1.50, "L6 - Director": 2.00,
}
RETENTION_RAISE_PCT = 0.08          # 8% targeted raise to watchlist employees
RAISE_RISK_REDUCTION = 0.45         # a retention raise cuts annual churn prob by 45%
PRODUCTIVITY_GAP_MONTHS = 3         # months of lost productivity per backfill (reactive)

print("=" * 64)
print("COMPENSATION COST & SCENARIO MODEL")
print("=" * 64)

scored = pd.read_csv(OUT / "scored_active_employees.csv")
scored["annual_churn_prob"] = scored["risk_score"] / 100.0
scored["repl_mult"] = scored["job_level"].map(REPLACEMENT_MULTIPLIER).fillna(1.0)
scored["replacement_cost"] = scored["base_salary"] * scored["repl_mult"]

# Horizon scaling: convert an annual churn probability to expected churn over 24m.
# P(leave within horizon) = 1 - (1 - p_annual)^(horizon/12)
horizon_years = HORIZON_MONTHS / 12
scored["p_leave_horizon"] = 1 - (1 - scored["annual_churn_prob"]).clip(0, 1) ** horizon_years

watch = scored[scored["on_watchlist"]].copy()

# ---------------------------------------------------------------------------
# SCENARIO A: STATUS QUO
#   Salary cost = current base over horizon. Expected replacement cost = sum of
#   p_leave_horizon * replacement_cost across the workforce.
# ---------------------------------------------------------------------------
base_salary_cost = scored["base_salary"].sum() * horizon_years
exp_replacement_A = (scored["p_leave_horizon"] * scored["replacement_cost"]).sum()
total_A = base_salary_cost + exp_replacement_A

# ---------------------------------------------------------------------------
# SCENARIO B: PROACTIVE RETENTION (raise the watchlist)
#   Extra salary = 8% raise on watchlist over horizon.
#   Their churn prob drops by RAISE_RISK_REDUCTION -> lower expected replacement.
# ---------------------------------------------------------------------------
extra_salary_B = (watch["base_salary"] * RETENTION_RAISE_PCT).sum() * horizon_years
# Watchlist new churn prob
watch_new_annual = watch["annual_churn_prob"] * (1 - RAISE_RISK_REDUCTION)
watch_new_horizon = 1 - (1 - watch_new_annual).clip(0, 1) ** horizon_years
watch_repl_new = (watch_new_horizon * watch["replacement_cost"]).sum()
# Non-watchlist unchanged
nonwatch = scored[~scored["on_watchlist"]]
nonwatch_repl = (nonwatch["p_leave_horizon"] * nonwatch["replacement_cost"]).sum()
exp_replacement_B = watch_repl_new + nonwatch_repl
total_B = base_salary_cost + extra_salary_B + exp_replacement_B

# ---------------------------------------------------------------------------
# SCENARIO C: REACTIVE / DO-NOTHING
#   No retention spend, full replacement cost, PLUS a productivity gap cost for
#   each expected backfill (3 months of salary lost to vacancy + ramp).
# ---------------------------------------------------------------------------
productivity_gap_C = (
    scored["p_leave_horizon"] * scored["base_salary"] * (PRODUCTIVITY_GAP_MONTHS / 12)
).sum()
exp_replacement_C = exp_replacement_A  # same churn as status quo
total_C = base_salary_cost + exp_replacement_C + productivity_gap_C

# ---------------------------------------------------------------------------
# Assemble tidy results
# ---------------------------------------------------------------------------
def eur(x): return round(float(x), 0)

summary = pd.DataFrame([
    {"scenario": "A. Status Quo",
     "base_salary_cost": eur(base_salary_cost),
     "retention_investment": 0.0,
     "expected_replacement_cost": eur(exp_replacement_A),
     "productivity_gap_cost": 0.0,
     "total_cost": eur(total_A)},
    {"scenario": "B. Proactive Retention",
     "base_salary_cost": eur(base_salary_cost),
     "retention_investment": eur(extra_salary_B),
     "expected_replacement_cost": eur(exp_replacement_B),
     "productivity_gap_cost": 0.0,
     "total_cost": eur(total_B)},
    {"scenario": "C. Reactive (Do Nothing)",
     "base_salary_cost": eur(base_salary_cost),
     "retention_investment": 0.0,
     "expected_replacement_cost": eur(exp_replacement_C),
     "productivity_gap_cost": eur(productivity_gap_C),
     "total_cost": eur(total_C)},
])
summary["vs_status_quo"] = (summary["total_cost"] - total_A).round(0)
summary.to_csv(OUT / "scenario_summary.csv", index=False)

net_benefit_B = total_A - total_B
print(f"\n{HORIZON_MONTHS}-month horizon | {len(scored)} active employees | "
      f"{len(watch)} on watchlist")
print(summary[["scenario", "retention_investment",
               "expected_replacement_cost", "total_cost", "vs_status_quo"]]
      .to_string(index=False))
print(f"\nProactive retention net saving vs status quo: EUR {net_benefit_B:,.0f}")
print(f"Reactive (do-nothing) extra cost vs status quo: "
      f"EUR {total_C - total_A:,.0f}")
print(f"Cost-of-inaction (C) minus cost-of-intervention (B): "
      f"EUR {total_C - total_B:,.0f}")

# ===========================================================================
# EXCEL EXPORT — fully formula-driven so finance can change assumptions live
# ===========================================================================
wb = Workbook()

# Styling helpers
HDR = Font(name="Arial", bold=True, size=12, color="FFFFFF")
SUB = Font(name="Arial", bold=True, size=10)
NORM = Font(name="Arial", size=10)
BLUE = Font(name="Arial", size=10, color="0000FF")   # inputs
BLACK = Font(name="Arial", size=10, color="000000")  # formulas
NAVY = PatternFill("solid", fgColor="1F3864")
LIGHT = PatternFill("solid", fgColor="D9E1F2")
YELLOW = PatternFill("solid", fgColor="FFFF00")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
EUR_FMT = '#,##0;(#,##0);-'
PCT_FMT = '0.0%'

# ---- Sheet 1: Assumptions (all blue inputs) ----
ws = wb.active
ws.title = "Assumptions"
ws["A1"] = "WORKFORCE COST SCENARIO MODEL — ASSUMPTIONS"
ws["A1"].font = HDR; ws["A1"].fill = NAVY
ws.merge_cells("A1:C1")
ws["A3"] = "Input"; ws["B3"] = "Value"; ws["C3"] = "Notes"
for c in "ABC": ws[f"{c}3"].font = SUB; ws[f"{c}3"].fill = LIGHT

assum = [
    ("Horizon (months)", HORIZON_MONTHS, "Projection window"),
    ("Retention raise %", RETENTION_RAISE_PCT, "Targeted raise to watchlist"),
    ("Risk reduction from raise", RAISE_RISK_REDUCTION,
     "Fractional cut to churn prob after raise"),
    ("Productivity gap (months)", PRODUCTIVITY_GAP_MONTHS,
     "Lost productivity per backfill (reactive)"),
]
r = 4
for name, val, note in assum:
    ws[f"A{r}"] = name; ws[f"A{r}"].font = NORM
    ws[f"B{r}"] = val; ws[f"B{r}"].font = BLUE; ws[f"B{r}"].fill = YELLOW
    if isinstance(val, float) and val < 1:
        ws[f"B{r}"].number_format = PCT_FMT
    ws[f"C{r}"] = note; ws[f"C{r}"].font = NORM
    r += 1
r += 1
ws[f"A{r}"] = "Replacement cost multiplier by job level (x annual salary):"
ws[f"A{r}"].font = SUB; r += 1
mult_start = r
for lvl, m in REPLACEMENT_MULTIPLIER.items():
    ws[f"A{r}"] = lvl; ws[f"A{r}"].font = NORM
    ws[f"B{r}"] = m; ws[f"B{r}"].font = BLUE; ws[f"B{r}"].fill = YELLOW
    ws[f"B{r}"].number_format = '0.00"x"'
    r += 1
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 46

# Named cells for formula references
HORIZON_CELL = "Assumptions!$B$4"
RAISE_CELL = "Assumptions!$B$5"
REDUCE_CELL = "Assumptions!$B$6"
PRODGAP_CELL = "Assumptions!$B$7"

# ---- Sheet 2: Employee detail (per-employee, formula-driven costs) ----
ws2 = wb.create_sheet("Employee Detail")
headers = ["Employee ID", "Department", "Job Level", "Base Salary (EUR)",
           "Risk Score", "Annual Churn Prob", "On Watchlist", "Repl. Multiplier",
           "Replacement Cost (EUR)", "P(leave 24m)",
           "Exp. Replacement (Status Quo)", "Exp. Replacement (Proactive)"]
for j, h in enumerate(headers, 1):
    c = ws2.cell(1, j, h); c.font = HDR; c.fill = NAVY
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
ws2.freeze_panes = "A2"

# Build a quick lookup of the multiplier cell row for each level
mult_row = {lvl: mult_start + i for i, lvl in enumerate(REPLACEMENT_MULTIPLIER)}

sd = scored.sort_values("risk_score", ascending=False).reset_index(drop=True)
for i, row in sd.iterrows():
    xr = i + 2
    ws2.cell(xr, 1, row["employee_id"]).font = NORM
    ws2.cell(xr, 2, row["department"]).font = NORM
    ws2.cell(xr, 3, row["job_level"]).font = NORM
    c = ws2.cell(xr, 4, round(row["base_salary"], 0)); c.font = BLUE; c.number_format = EUR_FMT
    c = ws2.cell(xr, 5, round(row["risk_score"], 1)); c.font = BLUE
    c = ws2.cell(xr, 6, f"=E{xr}/100"); c.font = BLACK; c.number_format = PCT_FMT
    ws2.cell(xr, 7, "Yes" if row["on_watchlist"] else "No").font = NORM
    # Replacement multiplier via VLOOKUP into Assumptions
    c = ws2.cell(xr, 8,
        f'=VLOOKUP(C{xr},Assumptions!$A${mult_start}:$B${mult_start+len(REPLACEMENT_MULTIPLIER)-1},2,FALSE)')
    c.font = BLACK; c.number_format = '0.00"x"'
    c = ws2.cell(xr, 9, f"=D{xr}*H{xr}"); c.font = BLACK; c.number_format = EUR_FMT
    # P(leave horizon) = 1-(1-annual)^(horizon/12)
    c = ws2.cell(xr, 10, f"=1-(1-F{xr})^({HORIZON_CELL}/12)")
    c.font = BLACK; c.number_format = PCT_FMT
    # Expected replacement status quo = P(leave)*replacement
    c = ws2.cell(xr, 11, f"=J{xr}*I{xr}"); c.font = BLACK; c.number_format = EUR_FMT
    # Expected replacement proactive: if watchlist, churn prob reduced
    c = ws2.cell(xr, 12,
        f'=IF(G{xr}="Yes",(1-(1-F{xr}*(1-{REDUCE_CELL}))^({HORIZON_CELL}/12))*I{xr},K{xr})')
    c.font = BLACK; c.number_format = EUR_FMT
    for col in range(1, 13):
        ws2.cell(xr, col).border = BORDER

last = len(sd) + 1
for col, w in zip("ABCDEFGHIJKL",
                  [13,16,13,15,10,13,11,12,16,12,18,18]):
    ws2.column_dimensions[col].width = w

# ---- Sheet 3: Scenario Summary (links to detail via SUM formulas) ----
ws3 = wb.create_sheet("Scenario Summary")
ws3["A1"] = "24-MONTH WORKFORCE COST — SCENARIO COMPARISON (EUR)"
ws3["A1"].font = HDR; ws3["A1"].fill = NAVY; ws3.merge_cells("A1:E1")

# Base salary cost over horizon (sum of base * horizon/12)
ws3["A3"] = "Component"; ws3["B3"] = "A. Status Quo"
ws3["C3"] = "B. Proactive Retention"; ws3["D3"] = "C. Reactive (Do Nothing)"
for col in "ABCD":
    ws3[f"{col}3"].font = SUB; ws3[f"{col}3"].fill = LIGHT
    ws3[f"{col}3"].alignment = Alignment(wrap_text=True, horizontal="center")

# Row 4: base salary cost (same across scenarios)
ws3["A4"] = "Base salary cost"
ws3["B4"] = f"=SUM('Employee Detail'!D2:D{last})*{HORIZON_CELL}/12"
ws3["C4"] = f"=B4+SUMIF('Employee Detail'!G2:G{last},\"Yes\",'Employee Detail'!D2:D{last})*{RAISE_CELL}*{HORIZON_CELL}/12"
ws3["D4"] = "=B4"
# Row 5: retention investment
ws3["A5"] = "Retention investment (raises)"
ws3["B5"] = 0
ws3["C5"] = f"=C4-B4"
ws3["D5"] = 0
# Row 6: expected replacement cost
ws3["A6"] = "Expected replacement cost"
ws3["B6"] = f"=SUM('Employee Detail'!K2:K{last})"
ws3["C6"] = f"=SUM('Employee Detail'!L2:L{last})"
ws3["D6"] = f"=B6"
# Row 7: productivity gap (reactive only)
ws3["A7"] = "Productivity gap cost"
ws3["B7"] = 0
ws3["C7"] = 0
ws3["D7"] = f"=SUMPRODUCT('Employee Detail'!J2:J{last},'Employee Detail'!D2:D{last})*{PRODGAP_CELL}/12"
# Row 8: TOTAL
ws3["A8"] = "TOTAL 24-MONTH COST"
ws3["B8"] = "=B4+B6+B7"
ws3["C8"] = "=C4+C6+C7"
ws3["D8"] = "=D4+D6+D7"
# Row 10: deltas
ws3["A10"] = "Cost vs Status Quo"
ws3["B10"] = "=B8-$B$8"
ws3["C10"] = "=C8-$B$8"
ws3["D10"] = "=D8-$B$8"
ws3["A12"] = "Net saving: Proactive vs Reactive"
ws3["B12"] = "=D8-C8"

for r_ in range(4, 9):
    ws3[f"A{r_}"].font = NORM if r_ != 8 else SUB
    for col in "BCD":
        cell = ws3[f"{col}{r_}"]
        cell.font = BLACK if r_ != 8 else Font(name="Arial", bold=True, size=10)
        cell.number_format = EUR_FMT
ws3["A8"].fill = LIGHT
for col in "BCD": ws3[f"{col}8"].fill = LIGHT
for r_ in [10, 12]:
    ws3[f"A{r_}"].font = SUB
    for col in "BCD":
        ws3[f"{col}{r_}"].number_format = EUR_FMT
        ws3[f"{col}{r_}"].font = BLACK
ws3.column_dimensions["A"].width = 30
for col in "BCD": ws3.column_dimensions[col].width = 22

# Insight note
ws3["A14"] = ("Read: Proactive retention spends on targeted raises but avoids most "
              "replacement cost; Reactive eats full replacement plus a productivity gap.")
ws3["A14"].font = Font(name="Arial", italic=True, size=9)
ws3.merge_cells("A14:D14")

xlsx_path = OUT / "compensation_scenario_model.xlsx"
wb.save(xlsx_path)
print(f"\nWrote Excel model -> {xlsx_path.name} "
      f"(Assumptions / Employee Detail / Scenario Summary, live formulas)")
